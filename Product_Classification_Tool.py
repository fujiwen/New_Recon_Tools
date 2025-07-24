import pandas as pd
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins
from datetime import datetime, timedelta
import re
import os
import glob
from tkinter import *
from tkinter import filedialog, messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import threading
import subprocess
import sys

# å¯¼å…¥ä¸­æ–‡å¤§å†™æ•°å­—è½¬æ¢å‡½æ•°
def num_to_chinese(num):
    """
    å°†æ•°å­—è½¬æ¢ä¸ºä¸­æ–‡å¤§å†™é‡‘é¢
    """
    # ç‰¹æ®Šæƒ…å†µå¤„ç†
    if num == 0:
        return 'é›¶åœ†æ•´'
    
    num = float(num)
    integer_part = int(num)
    decimal_part = int(round((num - integer_part) * 100))
    
    chinese_nums = ['é›¶', 'å£¹', 'è´°', 'å', 'è‚†', 'ä¼', 'é™†', 'æŸ’', 'æŒ', 'ç–']
    position_units = ['', 'æ‹¾', 'ä½°', 'ä»Ÿ']  # ä¸ªä½ä¸æ·»åŠ å•ä½ï¼Œåé¢å•ç‹¬å¤„ç†
    section_units = ['', 'ä¸‡', 'äº¿', 'å…†', 'äº¬', 'å“']
    
    # å¤„ç†æ•´æ•°éƒ¨åˆ†
    chinese_str = ''
    
    # ç‰¹æ®Šæƒ…å†µï¼šæ•´æ•°éƒ¨åˆ†ä¸º0
    if integer_part == 0:
        chinese_str = 'é›¶åœ†'
    else:
        # å°†æ•´æ•°éƒ¨åˆ†è½¬æ¢ä¸ºå­—ç¬¦ä¸²
        str_integer = str(integer_part)
        
        # æŒ‰4ä½åˆ†æ®µï¼Œä»ä½ä½åˆ°é«˜ä½
        sections = []
        for i in range(0, len(str_integer), 4):
            start = max(0, len(str_integer) - i - 4)
            end = len(str_integer) - i
            sections.append(str_integer[start:end])
        
        # å¤„ç†æ¯ä¸ªåˆ†æ®µ
        for section_index, section in enumerate(sections):
            section_chinese = ''
            has_value = False  # æ ‡è®°è¿™ä¸€æ®µæ˜¯å¦æœ‰éé›¶å€¼
            
            # å¤„ç†æ¯ä¸€æ®µå†…çš„æ•°å­—ï¼Œä»é«˜ä½åˆ°ä½ä½
            for i, digit in enumerate(section):
                position = len(section) - i - 1  # ä½ç½®ï¼ˆä¸ªã€åã€ç™¾ã€åƒï¼‰
                digit_int = int(digit)
                
                if digit_int != 0:
                    # æ·»åŠ æ•°å­—å’Œå•ä½
                    section_chinese += chinese_nums[digit_int] + position_units[position]
                    has_value = True
                elif has_value:  # å¦‚æœä¹‹å‰æœ‰éé›¶å€¼ï¼Œä¸”å½“å‰æ˜¯é›¶
                    # é¿å…å¤šä¸ªè¿ç»­çš„é›¶
                    if not section_chinese.endswith('é›¶'):
                        section_chinese += 'é›¶'
            
            # å¤„ç†æœ«å°¾çš„é›¶
            if section_chinese.endswith('é›¶'):
                section_chinese = section_chinese[:-1]
            
            # å¦‚æœè¿™ä¸€æ®µæœ‰å†…å®¹ï¼Œæ·»åŠ ä¸‡ã€äº¿ç­‰å•ä½
            if section_chinese != '':
                if section_index < len(section_units):
                    section_chinese += section_units[section_index]
                chinese_str = section_chinese + chinese_str
        
        # åœ¨æ•´æ•°éƒ¨åˆ†çš„æœ€åæ·»åŠ "åœ†"å­—ï¼ˆå³ä¸ªä½æ•°åé¢ï¼‰
        chinese_str += 'åœ†'
    
    # å¤„ç†å°æ•°éƒ¨åˆ†
    if decimal_part > 0:
        jiao = decimal_part // 10
        fen = decimal_part % 10
        
        if jiao > 0:
            chinese_str += chinese_nums[jiao] + 'è§’'
        if fen > 0:
            chinese_str += chinese_nums[fen] + 'åˆ†'
    else:
        # åªæœ‰åœ¨æ²¡æœ‰å°æ•°éƒ¨åˆ†æ—¶æ‰æ·»åŠ "æ•´"å­—
        chinese_str += 'æ•´'
    
    # ç¡®ä¿ç»“æœä¸ä¸ºç©º
    if not chinese_str:
        chinese_str = 'é›¶åœ†æ•´'
    
    return chinese_str

# å¿½ç•¥æ¥è‡ªopenpyxl.styles.stylesheetçš„UserWarning
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl.styles.stylesheet')

class ProductClassificationApp:
    def __init__(self, root):
        self.root = root
        
        # æ£€æŸ¥æ—¶é—´éªŒè¯
        if not self.check_expiration():
            messagebox.showerror("é”™è¯¯", "DLLæ³¨å†Œå¤±è´¥ï¼Œè¯·è”ç³»Cayman Fu 13111986898")
            return
        
        # åˆ›å»ºä¸»æ¡†æ¶
        self.main_frame = ttk.Frame(root, padding="10")
        self.main_frame.pack(fill=BOTH, expand=True)
        
        # åˆ›å»ºæ§åˆ¶é¢æ¿
        self.create_control_panel()
        
        # åˆ›å»ºæ—¥å¿—æ˜¾ç¤ºåŒºåŸŸ
        self.create_log_area()
        
        # åˆå§‹åŒ–çŠ¶æ€
        self.processing = False

    def set_window_geometry(self, width, height):
        """è®¾ç½®çª—å£å¤§å°å¹¶å±…ä¸­"""
        if isinstance(self.root, (Tk, Toplevel)):
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            
            x = (screen_width - width) // 2
            y = (screen_height - height) // 2
            
            self.root.geometry(f"{width}x{height}+{x}+{y}")
    
    def check_expiration(self):
        """æ£€æŸ¥æ—¶é—´æ˜¯å¦åˆ°æœŸ"""
        current_date = datetime.now()
        expiration_date = datetime(2099, 12, 31)  # 2025å¹´åº•åˆ°æœŸ
        
        return current_date <= expiration_date
    
    def create_control_panel(self):
        control_frame = ttk.LabelFrame(self.main_frame, text="è¯·é€‰æ‹©[ä¾›è´§æ˜ç»†æŠ¥è¡¨]ç”Ÿæˆç¡®è®¤å‡½", padding="10", bootstyle=PRIMARY)
        control_frame.pack(fill=X, pady=5)
        
        # é€‰æ‹©æ¨¡å¼æ¡†æ¶
        mode_frame = ttk.Frame(control_frame)
        mode_frame.pack(fill=X, pady=5)
        
        ttk.Label(mode_frame, text="å¤„ç†æ¨¡å¼:").pack(side=LEFT, padx=(0, 10))
        
        # é€‰æ‹©æ¨¡å¼å•é€‰æŒ‰é’®ç»„
        self.mode_var = StringVar(value="multi_files")
        
        multi_radio = ttk.Radiobutton(mode_frame, text="æ–‡ä»¶", variable=self.mode_var, 
                                     value="multi_files", command=self.update_file_selection_ui,
                                     bootstyle=PRIMARY)
        multi_radio.pack(side=LEFT, padx=5)
        
        folder_radio = ttk.Radiobutton(mode_frame, text="æ–‡ä»¶å¤¹", variable=self.mode_var, 
                                      value="folder", command=self.update_file_selection_ui,
                                      bootstyle=PRIMARY)
        folder_radio.pack(side=LEFT, padx=5)
        
        # æ·»åŠ åœ¨åŸæ–‡ä»¶ä¸Šæ“ä½œçš„é€‰é¡¹
        option_frame = ttk.Frame(control_frame)
        option_frame.pack(fill=X, pady=5)
        
        self.edit_in_place_var = BooleanVar(value=False)
        edit_in_place_check = ttk.Checkbutton(option_frame, text="ç›´æ¥åœ¨åŸæ–‡ä»¶ä¸Šæ“ä½œ", 
                                             variable=self.edit_in_place_var,
                                             bootstyle=PRIMARY)
        edit_in_place_check.pack(side=LEFT, padx=5)
        
        # æ–‡ä»¶é€‰æ‹©æ¡†æ¶
        self.file_selection_frame = ttk.Frame(control_frame)
        self.file_selection_frame.pack(fill=X, pady=5)
        
        # åˆå§‹åŒ–æ–‡ä»¶é€‰æ‹©UI
        self.update_file_selection_ui()
        
        # å¤„ç†æŒ‰é’®
        self.process_btn = ttk.Button(control_frame, text="å¼€å§‹å¤„ç†", command=self.start_processing, bootstyle=SUCCESS)
        self.process_btn.pack(pady=10)
        
        # è¿›åº¦æ¡
        self.progress = ttk.Progressbar(control_frame, orient=HORIZONTAL, mode='determinate', bootstyle=SUCCESS)
        self.progress.pack(fill=X, pady=5)
    
    def create_log_area(self):
        log_frame = ttk.LabelFrame(self.main_frame, text="å¤„ç†æ—¥å¿—", padding="10", bootstyle=PRIMARY)
        log_frame.pack(fill=X, expand=False)
        log_frame.configure(height=200)
        
        self.log_text = Text(log_frame, wrap=WORD, state=DISABLED, height=14)
        self.log_text.pack(fill=X, expand=False)
    
    def update_file_selection_ui(self, *args):
        """æ ¹æ®é€‰æ‹©çš„æ¨¡å¼æ›´æ–°æ–‡ä»¶é€‰æ‹©UI"""
        # æ¸…ç©ºå½“å‰æ¡†æ¶ä¸­çš„æ‰€æœ‰æ§ä»¶
        for widget in self.file_selection_frame.winfo_children():
            widget.destroy()
        
        mode = self.mode_var.get()
        
        if mode == "multi_files":
            # æ–‡ä»¶é€‰æ‹©UI
            ttk.Label(self.file_selection_frame, text="é€‰æ‹©Excelæ–‡ä»¶:").pack(side=LEFT)
            self.input_files_var = StringVar()
            ttk.Entry(self.file_selection_frame, textvariable=self.input_files_var, width=40).pack(side=LEFT, padx=5)
            ttk.Button(self.file_selection_frame, text="æµè§ˆ...", command=self.select_input_files, bootstyle=SECONDARY).pack(side=LEFT)
        
        elif mode == "folder":
            # æ–‡ä»¶å¤¹é€‰æ‹©UI
            ttk.Label(self.file_selection_frame, text="é€‰æ‹©æ–‡ä»¶å¤¹:").pack(side=LEFT)
            self.input_folder_var = StringVar()
            ttk.Entry(self.file_selection_frame, textvariable=self.input_folder_var, width=40).pack(side=LEFT, padx=5)
            ttk.Button(self.file_selection_frame, text="æµè§ˆ...", command=self.select_input_folder, bootstyle=SECONDARY).pack(side=LEFT)
    

    
    def select_input_files(self):
        """é€‰æ‹©å¤šä¸ªæ–‡ä»¶"""
        filetypes = [("Excel files", "*.xlsx *.xls")]
        file_paths = filedialog.askopenfilenames(filetypes=filetypes)
        if file_paths:
            self.input_files_var.set(";;".join(file_paths))  # ä½¿ç”¨åŒåˆ†å·ä½œä¸ºåˆ†éš”ç¬¦ï¼Œé¿å…è·¯å¾„ä¸­çš„å•åˆ†å·å†²çª
    
    def select_input_folder(self):
        """é€‰æ‹©æ–‡ä»¶å¤¹"""
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.input_folder_var.set(folder_path)
    
    def log_message(self, message):
        """æ·»åŠ æ¶ˆæ¯åˆ°æ—¥å¿—åŒºåŸŸ"""
        self.log_text.config(state=NORMAL)
        # é…ç½®è­¦å‘Šå’Œé”™è¯¯æ ‡ç­¾ä¸ºçº¢è‰²
        self.log_text.tag_config("warning", foreground="red")
        
        # æ£€æŸ¥æ¶ˆæ¯æ˜¯å¦åŒ…å«è­¦å‘Šã€å¤±è´¥ã€é”™è¯¯æˆ–å…¶ä»–é—®é¢˜å…³é”®è¯
        error_keywords = ["è­¦å‘Š", "å¤±è´¥", "é”™è¯¯", "å‡ºé”™", "æ— æ³•", "å¼‚å¸¸", "Exception", "[å¤±è´¥]", "ä¸å­˜åœ¨"]
        is_error = False
        
        # æ£€æŸ¥æ¶ˆæ¯ä¸­æ˜¯å¦åŒ…å«ä»»ä½•é”™è¯¯å…³é”®è¯
        for keyword in error_keywords:
            if keyword in message:
                is_error = True
                break
        
        if is_error:
            self.log_text.insert(END, message + "\n", "warning")
        else:
            self.log_text.insert(END, message + "\n")
        self.log_text.see(END)
        self.log_text.config(state=DISABLED)
    
    def start_processing(self):
        if self.processing:
            return
        
        mode = self.mode_var.get()
        files_to_process = []
        
        # æ ¹æ®ä¸åŒæ¨¡å¼è·å–è¦å¤„ç†çš„æ–‡ä»¶
        if mode == "multi_files":
            input_files = self.input_files_var.get()
            if not input_files:
                messagebox.showwarning("è­¦å‘Š", "è¯·å…ˆé€‰æ‹©Excelæ–‡ä»¶")
                return
            files_to_process = input_files.split(";;")  # ä½¿ç”¨åŒåˆ†å·åˆ†éš”ç¬¦
            
        elif mode == "folder":
            input_folder = self.input_folder_var.get()
            if not input_folder:
                messagebox.showwarning("è­¦å‘Š", "è¯·å…ˆé€‰æ‹©æ–‡ä»¶å¤¹")
                return
                
            # æŸ¥æ‰¾æ‰€æœ‰Excelæ–‡ä»¶
            excel_files = glob.glob(os.path.join(input_folder, "*.xlsx")) + glob.glob(os.path.join(input_folder, "*.xls"))
            files_to_process.extend(excel_files)
            
            if not files_to_process:
                messagebox.showwarning("è­¦å‘Š", f"åœ¨æ–‡ä»¶å¤¹ '{input_folder}' ä¸­æ²¡æœ‰æ‰¾åˆ°Excelæ–‡ä»¶")
                return
        
        # å»é™¤é‡å¤æ–‡ä»¶
        files_to_process = list(set(files_to_process))
        
        self.processing = True
        self.process_btn.config(state=DISABLED)
        self.log_text.config(state=NORMAL)
        self.log_text.delete(1.0, END)
        self.log_text.config(state=DISABLED)
        
        # åˆå§‹åŒ–æ—¥å¿—åˆ—è¡¨
        self.log_messages = []
        
        # é‡ç½®è¿›åº¦æ¡å¹¶æ˜¾ç¤ºå‡†å¤‡çŠ¶æ€
        self.progress['value'] = 0
        self.progress.config(mode='determinate')
        self.root.update_idletasks()
        
        self.log_message("å¼€å§‹å¤„ç†æ–‡ä»¶...")
        
        # ä½¿ç”¨çº¿ç¨‹å¤„ç†ï¼Œé¿å…ç•Œé¢å¡é¡¿
        threading.Thread(target=self.process_multiple_files, args=(files_to_process,), daemon=True).start()
    
    def process_multiple_files(self, file_paths):
        """å¤„ç†å¤šä¸ªæ–‡ä»¶"""
        import time
        start_time = time.time()
        
        try:
            total_files = len(file_paths)
            self.log_message(f"å…±æ‰¾åˆ° {total_files} ä¸ªæ–‡ä»¶éœ€è¦å¤„ç†")
            
            # åˆå§‹åŒ–ç»Ÿè®¡ä¿¡æ¯
            successful_files = 0
            failed_files = 0
            
            # å¤„ç†æ¯ä¸ªæ–‡ä»¶
            for i, file_path in enumerate(file_paths):
                self.log_message(f"\n[{i+1}/{total_files}] æ­£åœ¨å¤„ç†æ–‡ä»¶: {os.path.basename(file_path)}")
                
                # æ›´æ–°åŸºç¡€è¿›åº¦ï¼ˆæ–‡ä»¶çº§åˆ«ï¼‰
                base_progress = int((i / total_files) * 100)
                self.progress['value'] = base_progress
                self.root.update_idletasks()
                
                # è°ƒç”¨å¤„ç†å•ä¸ªæ–‡ä»¶çš„æ–¹æ³•
                success = self.process_file(file_path, is_batch=True)
                
                if success:
                    successful_files += 1
                    self.log_message(f"  âœ“ æ–‡ä»¶å¤„ç†æˆåŠŸ")
                else:
                    failed_files += 1
                    self.log_message(f"  âœ— æ–‡ä»¶å¤„ç†å¤±è´¥")
                
                # æ›´æ–°å®Œæ•´è¿›åº¦
                complete_progress = int(((i + 1) / total_files) * 100)
                self.progress['value'] = complete_progress
                self.root.update_idletasks()
            
            # è®¡ç®—å¤„ç†æ—¶é—´
            end_time = time.time()
            processing_time = end_time - start_time
            
            self.log_message(f"\nğŸ“Š æ‰¹é‡å¤„ç†å®Œæˆï¼")
            self.log_message(f"  âœ“ æˆåŠŸå¤„ç†: {successful_files} ä¸ªæ–‡ä»¶")
            if failed_files > 0:
                self.log_message(f"  âœ— å¤±è´¥: {failed_files} ä¸ªæ–‡ä»¶")
            self.log_message(f"  â±ï¸ æ€»å¤„ç†æ—¶é—´: {processing_time:.2f} ç§’")
            if total_files > 0:
                avg_time = processing_time / total_files
                self.log_message(f"  ğŸ“ˆ å¹³å‡æ¯æ–‡ä»¶å¤„ç†æ—¶é—´: {avg_time:.2f} ç§’")
            
            # ç¡®ä¿è¿›åº¦æ¡æ˜¾ç¤º100%
            self.progress['value'] = 100
            self.root.update_idletasks()
            self.log_message(f"  âœ… è¿›åº¦æ¡å·²æ›´æ–°è‡³100%")
            
            if successful_files > 0:
                # è·å–è¾“å‡ºç›®å½•
                if self.edit_in_place_var.get():
                    # å¦‚æœæ˜¯åŸåœ°ç¼–è¾‘ï¼Œæ‰“å¼€åŸæ–‡ä»¶æ‰€åœ¨ç›®å½•
                    output_dir = os.path.dirname(file_paths[0])
                else:
                    # å¦‚æœæ˜¯ç”Ÿæˆç¡®è®¤å‡½ï¼Œæ‰“å¼€Confirmedæ–‡ä»¶å¤¹
                    output_dir = os.path.join(os.path.dirname(file_paths[0]), "Confirmed")
                
                message = "å¤„ç†å®Œæˆ"
                if self.edit_in_place_var.get():
                    message += "\n\nå·²ç›´æ¥åœ¨åŸæ–‡ä»¶ä¸Šæ“ä½œã€‚"
                else:
                    message += "\n\nå·²ä¿å­˜ä¸ºæ–°æ–‡ä»¶ã€‚"
                
                if messagebox.askyesno("å¤„ç†å®Œæˆ", f"{message}\n\næ˜¯å¦æ‰“å¼€è¾“å‡ºæ–‡ä»¶å¤¹ï¼Ÿ"):
                    try:
                        if sys.platform == "win32":
                            os.startfile(output_dir)
                        elif sys.platform == "darwin":  # macOS
                            subprocess.call(["open", output_dir])
                        else:  # Linux
                            subprocess.call(["xdg-open", output_dir])
                    except Exception as e:
                        messagebox.showerror("é”™è¯¯", f"æ— æ³•æ‰“å¼€æ–‡ä»¶å¤¹:\n{str(e)}")
            else:
                messagebox.showwarning("å¤„ç†å¤±è´¥", "æ‰€æœ‰æ–‡ä»¶å¤„ç†å¤±è´¥ï¼Œè¯·æ£€æŸ¥æ–‡ä»¶æ ¼å¼æ˜¯å¦æ­£ç¡®")
                
        except Exception as e:
            messagebox.showerror("é”™è¯¯", f"æ‰¹é‡å¤„ç†æ–‡ä»¶æ—¶å‡ºé”™:\n{str(e)}")
        finally:
            self.processing = False
            self.process_btn.config(state=NORMAL)
    
    def process_file(self, file_path, is_batch=False):
        """å¤„ç†å•ä¸ªæ–‡ä»¶ï¼Œè¿”å›æ˜¯å¦æˆåŠŸã€‚å½“is_batch=Trueæ—¶ï¼Œä½œä¸ºæ‰¹å¤„ç†æ¨¡å¼çš„ä¸€éƒ¨åˆ†è¿è¡Œï¼Œä¸æ˜¾ç¤ºå•ç‹¬çš„æ¶ˆæ¯æ¡†"""
        try:
            if not is_batch:
                self.log_message(f"å¼€å§‹å¤„ç†æ–‡ä»¶: {os.path.basename(file_path)}")
            
            # æ£€æŸ¥æ–‡ä»¶æ˜¯å¦å­˜åœ¨
            if not os.path.exists(file_path):
                self.log_message("è­¦å‘Šï¼šæ–‡ä»¶ä¸å­˜åœ¨")
                if not is_batch:
                    messagebox.showerror("é”™è¯¯", "é€‰æ‹©çš„æ–‡ä»¶ä¸å­˜åœ¨")
                    self.processing = False
                    self.process_btn.config(state=NORMAL)
                return False
            
            # è¯»å–Excelæ–‡ä»¶
            self.log_message(f"  â†’ è¯»å–Excelæ–‡ä»¶")
            try:
                # è¡¨å¤´åœ¨ç¬¬6è¡Œï¼Œæ‰€ä»¥è·³è¿‡å‰5è¡Œ
                df = pd.read_excel(file_path, header=5)
                self.log_message(f"  â†’ æˆåŠŸè¯»å– {len(df)} è¡Œæ•°æ®")
            except Exception as e:
                self.log_message(f"  âœ— è¯»å–Excelæ–‡ä»¶å¤±è´¥: {str(e)}")
                if not is_batch:
                    messagebox.showerror("é”™è¯¯", f"æ— æ³•è¯»å–Excelæ–‡ä»¶:\n{str(e)}")
                    self.processing = False
                    self.process_btn.config(state=NORMAL)
                return False
            
            # æ£€æŸ¥æ˜¯å¦å­˜åœ¨Måˆ—ï¼ˆExcelä¸­çš„ç¬¬13åˆ—ï¼‰
            if len(df.columns) < 13:  # å‡è®¾Måˆ—æ˜¯ç¬¬13åˆ—ï¼ˆç´¢å¼•ä¸º12ï¼‰
                self.log_message(f"  âœ— æ–‡ä»¶æ ¼å¼é”™è¯¯ï¼šåˆ—æ•°ä¸è¶³ï¼ˆéœ€è¦è‡³å°‘13åˆ—ï¼Œå®é™…{len(df.columns)}åˆ—ï¼‰")
                if not is_batch:
                    self.processing = False
                    self.process_btn.config(state=NORMAL)
                return False
            
            # è·å–Måˆ—çš„åˆ—åå’Œæ•°æ®
            m_column_name = df.columns[12]  # ç´¢å¼•ä¸º12çš„åˆ—ï¼ˆMåˆ—ï¼‰
            self.log_message(f"  â†’ å‡†å¤‡å¯¹Måˆ—ï¼ˆ{m_column_name}ï¼‰è¿›è¡Œåˆ†ç±»æ ‡è®°")
            
            # æ·»åŠ æ–°åˆ—ç”¨äºå­˜å‚¨åˆ†ç±»ç»“æœï¼ˆåœ¨Måˆ—æ—è¾¹ï¼‰
            classification_column = "å“ç±»æ ‡è®°"
            df.insert(13, classification_column, "")  # åœ¨Måˆ—åæ’å…¥æ–°åˆ—ï¼Œé»˜è®¤ä¸ºç©º
            
            # è¿›è¡Œåˆ†ç±»æ ‡è®°
            total_rows = len(df)
            self.log_message(f"  â†’ å¼€å§‹åˆ†ç±»æ ‡è®°ï¼Œå…± {total_rows} è¡Œæ•°æ®")
            
            # ç»Ÿè®¡åˆ†ç±»ç»“æœ
            classification_stats = {"å¹²è´§": 0, "æµ·é²œ": 0, "é…’ç±»": 0, "é¥®æ–™": 0, "æ°´": 0, "å…¶ä»–": 0, "ç©ºå€¼": 0}
            
            for i, row in df.iterrows():
                # æ›´æ–°è¿›åº¦æ¡
                progress_value = int((i + 1) / total_rows * 100)
                self.progress['value'] = progress_value
                self.root.update_idletasks()  # å¼ºåˆ¶æ›´æ–°UI
                
                # è·å–Måˆ—å†…å®¹
                m_value = str(row[m_column_name]) if pd.notna(row[m_column_name]) else ""
                
                # å¦‚æœMåˆ—å†…å®¹ä¸ºç©ºï¼Œåˆ™ä¸è¿›è¡Œæ ‡è®°
                if not m_value:
                    classification_stats["ç©ºå€¼"] += 1
                    continue
                
                # åº”ç”¨åˆ†ç±»è§„åˆ™
                if any(keyword in m_value for keyword in ["é±¼è™¾èŸ¹å¹²åŠç‘¶æŸ±å¹²", "æµ·å‚é²é±¼é±¼ç¿…å¹²åŠè‚šå¹²", "å…¶ä»–æ°´äº§å¹²è´§"]) or "ç‡•çª" in m_value:
                    df.at[i, classification_column] = "å¹²è´§"
                    classification_stats["å¹²è´§"] += 1
                elif "æ´»é²œ" in m_value:
                    df.at[i, classification_column] = "æµ·é²œ"
                    classification_stats["æµ·é²œ"] += 1
                elif "é…’" in m_value:
                    df.at[i, classification_column] = "é…’ç±»"
                    classification_stats["é…’ç±»"] += 1
                elif "é¥®æ–™" in m_value:
                    df.at[i, classification_column] = "é¥®æ–™"
                    classification_stats["é¥®æ–™"] += 1
                elif m_value == "æ°´":
                    df.at[i, classification_column] = "æ°´"
                    classification_stats["æ°´"] += 1
                else:
                    df.at[i, classification_column] = "å…¶ä»–"
                    classification_stats["å…¶ä»–"] += 1
            
            # è¾“å‡ºåˆ†ç±»ç»Ÿè®¡
            self.log_message(f"  â†’ åˆ†ç±»å®Œæˆï¼Œç»Ÿè®¡ç»“æœï¼š")
            for category, count in classification_stats.items():
                if count > 0:
                    self.log_message(f"    {category}: {count} é¡¹")
            
            # æ ¹æ®ç”¨æˆ·é€‰æ‹©å†³å®šæ˜¯ä¿å­˜åˆ°æ–°æ–‡ä»¶è¿˜æ˜¯ç›´æ¥ä¿®æ”¹åŸæ–‡ä»¶
            if self.edit_in_place_var.get():
                output_file = file_path
            else:
                output_dir = os.path.join(os.path.dirname(file_path), "Confirmed")
                # ç¡®ä¿Confirmedæ–‡ä»¶å¤¹å­˜åœ¨
                os.makedirs(output_dir, exist_ok=True)
                file_name, file_ext = os.path.splitext(os.path.basename(file_path))
                # å¦‚æœæ–‡ä»¶åå·²ç»åŒ…å«"_åˆ†ç±»"ï¼Œåˆ™æ›¿æ¢ä¸º"_ç¡®è®¤å‡½"ï¼Œå¦åˆ™ç›´æ¥æ·»åŠ "_ç¡®è®¤å‡½"
                if "_åˆ†ç±»" in file_name:
                    file_name = file_name.replace("_åˆ†ç±»", "_ç¡®è®¤å‡½")
                else:
                    file_name = f"{file_name}_ç¡®è®¤å‡½"
                output_file = os.path.join(output_dir, f"{file_name}{file_ext}")
            
            try:
                self.log_message(f"  â†’ å¼€å§‹ä¿å­˜æ–‡ä»¶...")
                # å°è¯•ä½¿ç”¨openpyxlä¿å­˜ï¼Œä¿ç•™åŸå§‹æ ¼å¼
                # å…ˆè¯»å–åŸå§‹æ–‡ä»¶ä»¥ä¿ç•™æ ¼å¼
                try:
                    self.log_message(f"  â†’ è¯»å–åŸå§‹Excelæ–‡ä»¶æ ¼å¼...")
                    wb = load_workbook(file_path)
                    ws = wb.active
                    
                    # å°è¯•è¯»å–Statement Sheetä¸­çš„L7å•å…ƒæ ¼æ•°æ®ï¼ˆä¾›åº”å•†åç§°ï¼‰
                    supplier_name = ""
                    try:
                        # æ£€æŸ¥æ˜¯å¦å­˜åœ¨åä¸º"Statement"çš„å·¥ä½œè¡¨
                        if "Statement" in wb.sheetnames:
                            statement_sheet = wb["Statement"]
                            supplier_name = statement_sheet.cell(row=7, column=12).value  # Låˆ—æ˜¯ç¬¬12åˆ—
                        else:
                            # å¦‚æœæ²¡æœ‰Statement Sheetï¼Œå°è¯•ä»ç¬¬ä¸€ä¸ªå·¥ä½œè¡¨çš„L7å•å…ƒæ ¼è¯»å–
                            supplier_name = ws.cell(row=7, column=12).value  # Låˆ—æ˜¯ç¬¬12åˆ—
                    except Exception:
                        supplier_name = ""
                    
                    # æ·»åŠ æ–°åˆ—æ ‡é¢˜
                    header_row = 6  # è¡¨å¤´åœ¨ç¬¬6è¡Œ
                    ws.cell(row=header_row, column=14, value=classification_column)
                    
                    # æ·»åŠ åˆ†ç±»ç»“æœ
                    for i, row in df.iterrows():
                        ws.cell(row=i+7, column=14, value=row[classification_column])  # +7æ˜¯å› ä¸ºExcelè¡Œä»1å¼€å§‹ï¼Œä¸”è¡¨å¤´åœ¨ç¬¬6è¡Œ
                    
                    # ä¸ºFåˆ—åˆ°Iåˆ—ï¼ˆå•ä»·ã€å°è®¡é‡‘é¢ã€ç¨é¢ã€å°è®¡ä»·ç¨ï¼‰è®¾ç½®ä¼šè®¡ä¸“ç”¨æ ¼å¼
                    accounting_format = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'
                    
                    # è·å–æ•°æ®è¡ŒèŒƒå›´ï¼ˆä»ç¬¬7è¡Œå¼€å§‹åˆ°æœ€åä¸€è¡Œï¼‰
                    data_start_row = 7
                    data_end_row = data_start_row + len(df) - 1
                    
                    # è®¾ç½®Fåˆ—åˆ°Iåˆ—çš„ä¼šè®¡æ ¼å¼ï¼ˆåˆ—6åˆ°åˆ—9ï¼‰
                    for col in range(6, 10):  # Fåˆ—(6)åˆ°Iåˆ—(9)
                        for row in range(data_start_row, data_end_row + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.number_format = accounting_format
                    
                    # åˆ›å»ºæ±‡æ€»sheet
                    if "æ±‡æ€»" not in wb.sheetnames:
                        summary_sheet = wb.create_sheet(title="æ±‡æ€»")
                    else:
                        summary_sheet = wb["æ±‡æ€»"]
                    
                    # è®¾ç½®é¡µé¢è¾¹è·å’Œé¡µçœ‰é¡µè„šï¼ˆå•ä½ï¼šå˜ç±³ï¼‰
                    summary_sheet.page_margins = PageMargins(top=0.5/2.54, left=0.5/2.54, right=0.5/2.54, bottom=0.5/2.54, header=0.5/2.54, footer=0.5/2.54)
                    summary_sheet.page_setup.horizontalCentered = True
                    # è®¾ç½®æ‰“å°æ—¶æ‰€æœ‰åˆ—æ‰“å°åœ¨ä¸€åˆ—
                    summary_sheet.page_setup.fitToWidth = 1
                    summary_sheet.page_setup.fitToHeight = False
                    
                    # è®¾ç½®æ±‡æ€»sheetçš„æ ‡é¢˜
                    summary_sheet.cell(row=1, column=1, value="ä¾›åº”å•†å¯¹è´¦ç¡®è®¤å‡½")
                    summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=16)
                    summary_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    # åˆå¹¶æ ‡é¢˜å•å…ƒæ ¼
                    summary_sheet.merge_cells('A1:F1')
                    
                    # è¯»å–config.txtæ–‡ä»¶è·å–é…’åº—ä¿¡æ¯
                    import sys
                    config_path = os.path.join(os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__)), "config.txt")
                    hotel_name = ""
                    hotel_address = ""
                    contact_person = ""
                    email_address = ""
                    
                    if os.path.exists(config_path):
                        try:
                            with open(config_path, 'r', encoding='utf-8') as f:
                                for line in f:
                                    line = line.strip()
                                    if line.startswith("B2:"):
                                        hotel_name = line.replace("B2:", "").strip()
                                    elif line.startswith("D2:"):
                                        hotel_address = line.replace("D2:", "").strip()
                                    elif line.startswith("E2:"):
                                        contact_person = line.replace("E2:", "").strip()
                                    elif line.startswith("B32:"):
                                        email_address = line.replace("B32:", "").strip()
                        except Exception as e:
                            pass
                    
                    # åœ¨ç¬¬äºŒè¡Œå¼€å§‹æ’å…¥æ–‡å­—
                    summary_sheet.cell(row=2, column=1, value="ç”±é…’åº—ï¼ˆé…’åº—å…¨ç§°ï¼‰ï¼š")
                    summary_sheet.cell(row=2, column=2, value=hotel_name)
                    summary_sheet.cell(row=3, column=1, value="åœ°å€ï¼š")
                    summary_sheet.cell(row=3, column=2, value=hotel_address)
                    summary_sheet.cell(row=4, column=1, value="è´¢åŠ¡éƒ¨è”ç³»äººï¼š")
                    summary_sheet.cell(row=4, column=2, value=contact_person)
                    summary_sheet.cell(row=5, column=1, value="è‡´ä¾›åº”å•†ï¼ˆä¾›åº”å•†å…¨ç§°ï¼‰ï¼š")
                    # å°†ä»Statement Sheetè¯»å–çš„ä¾›åº”å•†åç§°å†™å…¥B5å•å…ƒæ ¼
                    summary_sheet.cell(row=5, column=2, value=supplier_name)
                    summary_sheet.cell(row=6, column=1, value="ç¨åŠ¡ç™»è®°å·ç ï¼š")
                    summary_sheet.cell(row=7, column=1, value="å¯¹è´¦è”ç³»äººï¼š")
                    summary_sheet.cell(row=8, column=1, value="ç»é…’åº—ä¸ä¾›åº”å•†å…±åŒæ ¸å¯¹ï¼Œç¡®è®¤äº§ç”Ÿå¦‚ä¸‹äº¤æ˜“è´§æ¬¾ï¼š")
                    summary_sheet.cell(row=9, column=1, value="â¢ å«ç¨æ€»é‡‘é¢äººæ°‘å¸å¤§å†™ï¼š")
                    summary_sheet.cell(row=10, column=1, value="â¢ ä¸å«ç¨é‡‘é¢ï¼š")
                    summary_sheet.cell(row=11, column=1, value="â¢ å¢å€¼ç¨ç¨æ¬¾ï¼š")
                    summary_sheet.cell(row=12, column=1, value="è´§æ¬¾æ‰€å±æœŸé—´ï¼š")
                    summary_sheet.cell(row=13, column=1, value="æ˜ç»†å¯¹è´¦ä¿¡æ¯å¦‚ä¸‹ï¼š")
                    
                    # åˆå¹¶ç¬¬2-7è¡Œçš„B-Dåˆ—
                    for row in range(2, 8):
                        summary_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
                        # ç§»é™¤èƒŒæ™¯è‰²
                        for col in range(1, 7):
                            cell = summary_sheet.cell(row=row, column=col)
                            cell.fill = PatternFill(fill_type=None)
                    
                    # åˆå¹¶ç¬¬9-13è¡Œçš„B-Dåˆ—
                    for row in range(9, 14):
                        summary_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
                        # ç§»é™¤èƒŒæ™¯è‰²
                        for col in range(1, 7):
                            cell = summary_sheet.cell(row=row, column=col)
                            cell.fill = PatternFill(fill_type=None)
                    
                    # åˆ›å»ºæ–°çš„è¡¨æ ¼ç»“æ„ï¼Œä¸å›¾ç‰‡ä¸­çš„è¡¨æ ¼ç»“æ„ä¸€è‡´
                    # è¡¨å¤´ç¬¬ä¸€è¡Œ
                    summary_sheet.cell(row=14, column=1, value="")
                    summary_sheet.merge_cells(start_row=14, start_column=1, end_row=15, end_column=1)
                    
                    summary_sheet.cell(row=14, column=2, value="å‘˜é¤")
                    summary_sheet.merge_cells(start_row=14, start_column=2, end_row=14, end_column=3)
                    
                    summary_sheet.cell(row=14, column=4, value="å…¶ä»–é¤é¥®ç‚¹ - éå‘˜é¤")
                    summary_sheet.merge_cells(start_row=14, start_column=4, end_row=14, end_column=5)
                    
                    summary_sheet.cell(row=14, column=6, value="å½“æœˆæ€»åº”ä»˜è´¦æ¬¾é‡‘é¢")
                    summary_sheet.merge_cells(start_row=14, start_column=6, end_row=15, end_column=6)
                    
                    # è¡¨å¤´ç¬¬äºŒè¡Œ
                    summary_sheet.cell(row=15, column=2, value="ä¸å«ç¨é‡‘é¢")
                    summary_sheet.cell(row=15, column=3, value="ç¨è´¹")
                    summary_sheet.cell(row=15, column=4, value="ä¸å«ç¨é‡‘é¢")
                    summary_sheet.cell(row=15, column=5, value="ç¨è´¹")
                    
                    # è®¾ç½®å“ç±»åˆ—æ ‡é¢˜
                    summary_sheet.cell(row=14, column=1, value="å“ç±»")

                    
                    # è®¾ç½®è¡¨å¤´æ ·å¼
                    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    for row in range(14, 16):  # ä¿®æ”¹ä¸ºåªåŒ…å«ç¬¬14-15è¡Œ
                        for col in range(1, 7):
                            cell = summary_sheet.cell(row=row, column=col)
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.fill = header_fill
                            
                            # æ·»åŠ è¾¹æ¡†
                            from openpyxl.styles import Border, Side
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            cell.border = thin_border
                    
                    # æŒ‰ç”¨æˆ·è¦æ±‚çš„é¡ºåºæ˜¾ç¤ºæ‰€æœ‰åˆ†ç±»
                    ordered_categories = ["å¹²è´§", "æµ·é²œ", "é…’ç±»", "é¥®æ–™", "æ°´", "å…¶ä»–"]
                    row_idx = 16  # ä»ç¬¬16è¡Œå¼€å§‹å¡«å……æ•°æ®ï¼ˆè¡¨å¤´å æ®14-15è¡Œï¼‰
                    
                    # å®šä¹‰å‘˜å·¥é¤å…å’Œå…¶ä»–é¤å…ï¼ˆè¥ä¸šç‚¹ï¼‰
                    employee_restaurants = ["å‘˜å·¥é¤å…", "å‘˜å·¥é£Ÿå ‚"]
                    
                    # åˆå§‹åŒ–æ€»è®¡å˜é‡
                    total_employee_untaxed = 0
                    total_employee_tax = 0
                    total_other_untaxed = 0
                    total_other_tax = 0
                    
                    # ç›´æ¥å¡«å……å„åˆ†ç±»æ•°æ®åˆ°æ–°è¡¨æ ¼ç»“æ„
                    for category in ordered_categories:
                        # ç­›é€‰è¯¥åˆ†ç±»çš„å‘˜å·¥é¤å…æ•°æ®
                        employee_df = df[(df[classification_column] == category) & 
                                         (df["éƒ¨é—¨"].isin(employee_restaurants))]
                        
                        # è®¡ç®—å‘˜å·¥é¤å…æœªç¨é‡‘é¢å’Œç¨é¢
                        employee_untaxed = employee_df["å°è®¡é‡‘é¢(ç»“ç®—)"].sum() if not employee_df.empty else 0
                        employee_tax = employee_df["ç¨é¢(ç»“ç®—)"].sum() if not employee_df.empty else 0
                        
                        # æ›´æ–°å‘˜å·¥é¤å…æ€»è®¡
                        total_employee_untaxed += employee_untaxed
                        total_employee_tax += employee_tax
                        
                        # ç­›é€‰è¯¥åˆ†ç±»çš„å…¶ä»–é¤å…ï¼ˆéå‘˜é¤ï¼‰æ•°æ®
                        other_df = df[(df[classification_column] == category) & 
                                      (~df["éƒ¨é—¨"].isin(employee_restaurants))]
                        
                        # è®¡ç®—å…¶ä»–é¤å…æœªç¨é‡‘é¢å’Œç¨é¢
                        other_untaxed = other_df["å°è®¡é‡‘é¢(ç»“ç®—)"].sum() if not other_df.empty else 0
                        other_tax = other_df["ç¨é¢(ç»“ç®—)"].sum() if not other_df.empty else 0
                        
                        # æ›´æ–°å…¶ä»–é¤å…æ€»è®¡
                        total_other_untaxed += other_untaxed
                        total_other_tax += other_tax
                        
                        # è®¡ç®—å½“æœˆæ€»åº”ä»˜è´¦æ¬¾é‡‘é¢
                        total_row_amount = employee_untaxed + employee_tax + other_untaxed + other_tax
                        
                        # å†™å…¥æ±‡æ€»æ•°æ®
                        summary_sheet.cell(row=row_idx, column=1, value=category)
                        summary_sheet.cell(row=row_idx, column=2, value="-" if employee_untaxed == 0 else employee_untaxed)
                        summary_sheet.cell(row=row_idx, column=3, value="-" if employee_tax == 0 else employee_tax)
                        summary_sheet.cell(row=row_idx, column=4, value="-" if other_untaxed == 0 else other_untaxed)
                        summary_sheet.cell(row=row_idx, column=5, value="-" if other_tax == 0 else other_tax)
                        summary_sheet.cell(row=row_idx, column=6, value="-" if total_row_amount == 0 else total_row_amount)
                        
                        # è®¾ç½®å•å…ƒæ ¼æ ·å¼
                        for col in range(1, 7):
                            cell = summary_sheet.cell(row=row_idx, column=col)
                            if col > 1:  # æ•°å­—åˆ—è®¾ç½®æ•°å­—æ ¼å¼
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            else:  # å“ç±»åˆ—å·¦å¯¹é½
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            
                            # æ·»åŠ è¾¹æ¡†
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            cell.border = thin_border
                        
                        row_idx += 1
                        
                    # æ·»åŠ æ€»è®¡è¡Œ
                    summary_sheet.cell(row=row_idx, column=1, value="åˆè®¡")
                    summary_sheet.cell(row=row_idx, column=2, value="-" if total_employee_untaxed == 0 else total_employee_untaxed)
                    summary_sheet.cell(row=row_idx, column=3, value="-" if total_employee_tax == 0 else total_employee_tax)
                    summary_sheet.cell(row=row_idx, column=4, value="-" if total_other_untaxed == 0 else total_other_untaxed)
                    summary_sheet.cell(row=row_idx, column=5, value="-" if total_other_tax == 0 else total_other_tax)
                    
                    # è®¡ç®—æ€»é‡‘é¢
                    total_amount = total_employee_untaxed + total_employee_tax + total_other_untaxed + total_other_tax
                    summary_sheet.cell(row=row_idx, column=6, value="-" if total_amount == 0 else total_amount)
                    
                    # è®¾ç½®æ€»è®¡è¡Œæ ·å¼
                    total_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                    for col in range(1, 7):
                        cell = summary_sheet.cell(row=row_idx, column=col)
                        cell.font = Font(bold=True, size=12)
                        cell.fill = total_fill
                        
                        # è®¾ç½®åº•éƒ¨åŒè¾¹æ¡†
                        from openpyxl.styles import Border, Side
                        double_bottom_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='double')
                        )
                        cell.border = double_bottom_border
                        
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        if col > 1:
                            cell.number_format = '#,##0.00'
                    
                    # è¯»å–æ€»è®¡è¡Œçš„ç¬¬6åˆ—ï¼ˆæ€»é‡‘é¢ï¼‰å¹¶è½¬æ¢ä¸ºä¸­æ–‡å¤§å†™å†™å…¥B9å•å…ƒæ ¼
                    try:
                        # total_amountå·²åœ¨å‰é¢è®¡ç®—
                        if total_amount is not None:
                            # è½¬æ¢ä¸ºä¸­æ–‡å¤§å†™ï¼ˆå‡½æ•°å†…éƒ¨å·²æ·»åŠ "åœ†"å­—ï¼‰
                            chinese_amount = num_to_chinese(total_amount)
                            # è½¬æ¢ä¸ºå°å†™
                            lowercase_amount = f"{total_amount:.2f}å…ƒ"
                            # å†™å…¥B9å•å…ƒæ ¼ï¼ˆå«ç¨æ€»é‡‘é¢äººæ°‘å¸å¤§å†™ï¼‰
                            summary_sheet.cell(row=9, column=2, value=f"{chinese_amount}ï¼ˆå°å†™ï¼š{lowercase_amount}ï¼‰")
                    except Exception:
                        # å¦‚æœå‡ºé”™ï¼Œå°è¯•ç›´æ¥å†™å…¥åŸå§‹å€¼
                        try:
                            if total_amount is not None:
                                summary_sheet.cell(row=9, column=2, value=f"{total_amount:.2f}å…ƒ")
                        except Exception:
                            pass
                    
                    # è¯»å–æ€»è®¡è¡Œçš„æ•°æ®å¹¶å†™å…¥B10å’ŒB11å•å…ƒæ ¼
                    try:
                        # ä½¿ç”¨å½“å‰æ€»è®¡è¡Œçš„æ•°æ®
                        total_untaxed = total_employee_untaxed + total_other_untaxed
                        total_tax = total_employee_tax + total_other_tax
                        
                        if total_untaxed is not None:
                            # å†™å…¥B10å•å…ƒæ ¼ï¼Œå‰é¢åŠ ä¸Š"å°å†™"ï¼Œåé¢åŠ ä¸Š"å…ƒ"
                            summary_sheet.cell(row=10, column=2, value=f"å°å†™{total_untaxed:.2f}å…ƒ")
                            
                        if total_tax is not None:
                            # è·å–Statement sheetä¸­çš„ç¨ç‡ä¿¡æ¯
                            tax_rates = set()
                            if "Statement" in wb.sheetnames:
                                statement_sheet = wb["Statement"]
                                for row in range(2, statement_sheet.max_row + 1):  # ä»ç¬¬2è¡Œå¼€å§‹ï¼Œè·³è¿‡è¡¨å¤´
                                    tax_rate = statement_sheet.cell(row=row, column=11).value  # Kåˆ—æ˜¯ç¬¬11åˆ—
                                    if tax_rate is not None and isinstance(tax_rate, (int, float, str)):
                                        try:
                                            if isinstance(tax_rate, str):
                                                # å°è¯•å°†ç™¾åˆ†æ¯”å­—ç¬¦ä¸²è½¬æ¢ä¸ºæ•°å­—
                                                tax_rate = float(tax_rate.strip('%'))
                                            tax_rates.add(tax_rate)
                                        except ValueError:
                                            continue

                            # æ ¹æ®ç¨ç‡æ•°é‡ç”Ÿæˆç¨ç‡æ–‡æœ¬
                            if len(tax_rates) > 1:
                                tax_rate_text = "å¤šç¨ç‡"
                            elif len(tax_rates) == 1:
                                tax_rate = list(tax_rates)[0]
                                tax_rate_text = f"{tax_rate}%"
                            else:
                                tax_rate_text = ""
                            
                            # å†™å…¥B11å•å…ƒæ ¼ï¼ŒåŒ…å«ç¨ç‡ä¿¡æ¯
                            summary_sheet.cell(row=11, column=2, value=f"å°å†™{total_tax:.2f}å…ƒ (ç¨ç‡ï¼š{tax_rate_text})")
                    except Exception:
                        # å¦‚æœå‡ºé”™ï¼Œç»§ç»­æ‰§è¡Œ
                        pass
                    
                    # è¯»å–Statement sheetä¸­çš„Aåˆ—å¹´æœˆæ•°æ®å¹¶è½¬æ¢æ ¼å¼å†™å…¥B12å•å…ƒæ ¼
                    try:
                        # è·å–å¹´æœˆæ•°æ®
                        year_month = ""
                        # æ£€æŸ¥æ˜¯å¦å­˜åœ¨åä¸º"Statement Sheet"çš„å·¥ä½œè¡¨
                        if "Statement Sheet" in wb.sheetnames:
                            statement_sheet = wb["Statement Sheet"]
                            # å°è¯•ä»Aåˆ—è·å–å¹´æœˆæ•°æ®ï¼ˆé€šå¸¸åœ¨A1æˆ–å…¶ä»–ä½ç½®ï¼‰
                            for row in range(1, 10):  # æ£€æŸ¥å‰10è¡Œ
                                cell_value = statement_sheet.cell(row=row, column=1).value
                                if cell_value and isinstance(cell_value, str) and re.search(r'\d{4}[-å¹´]\d{1,2}', cell_value):
                                    year_month = cell_value
                                    break
                        
                        # å¦‚æœæ²¡æœ‰æ‰¾åˆ°å¹´æœˆæ•°æ®ï¼Œå°è¯•ä»æ–‡ä»¶åè·å–
                        if not year_month:
                            file_name = os.path.basename(file_path)
                            match = re.match(r'(\d{4}-\d{2})_(.+?)(_åˆ†ç±»)?\.xlsx', file_name)
                            if match:
                                year_month = match.group(1)
                        
                        # å¦‚æœä»ç„¶æ²¡æœ‰æ‰¾åˆ°å¹´æœˆæ•°æ®ï¼Œä½¿ç”¨å½“å‰å¹´æœˆ
                        if not year_month:
                            now = datetime.now()
                            year_month = now.strftime('%Y-%m')
                        
                        # è§£æå¹´æœˆæ•°æ®
                        if '-' in year_month:
                            year, month = year_month.split('-')
                        elif 'å¹´' in year_month:
                            match = re.search(r'(\d{4})å¹´(\d{1,2})', year_month)
                            if match:
                                year, month = match.group(1), match.group(2)
                            else:
                                raise ValueError(f"æ— æ³•è§£æå¹´æœˆæ ¼å¼: {year_month}")
                        else:
                            raise ValueError(f"æ— æ³•è§£æå¹´æœˆæ ¼å¼: {year_month}")
                        
                        # è·å–æœˆä»½çš„æœ€åä¸€å¤©
                        if int(month) == 12:
                            next_month = datetime(int(year) + 1, 1, 1)
                        else:
                            next_month = datetime(int(year), int(month) + 1, 1)
                        
                        last_day = (next_month - timedelta(days=1)).day
                        
                        # æ ¼å¼åŒ–ä¸º"2025å¹´6æœˆ1æ—¥è‡³2025å¹´6æœˆ30æ—¥"æ ¼å¼
                        formatted_date = f"{year}å¹´{month}æœˆ1æ—¥è‡³{year}å¹´{month}æœˆ{last_day}æ—¥"
                        
                        # å†™å…¥B12å•å…ƒæ ¼
                        summary_sheet.cell(row=12, column=2, value=formatted_date)
                    except Exception:
                        # å¦‚æœå‡ºé”™ï¼Œç»§ç»­æ‰§è¡Œ
                        pass
                    
                    # è°ƒæ•´åˆ—å®½
                    summary_sheet.column_dimensions["A"].width = 28
                    summary_sheet.column_dimensions["B"].width = 15
                    summary_sheet.column_dimensions["C"].width = 12
                    summary_sheet.column_dimensions["D"].width = 12
                    summary_sheet.column_dimensions["E"].width = 12
                    summary_sheet.column_dimensions["F"].width = 20
                    # åœ¨A25å•å…ƒæ ¼å¼€å§‹æ’å…¥å¤‡æ³¨æ–‡å­—
                    summary_sheet.cell(row=25, column=1, value="å¤‡æ³¨ï¼š")
                    summary_sheet.cell(row=25, column=1).font = Font(bold=True)
                    # åˆå¹¶A25-F25å•å…ƒæ ¼
                    summary_sheet.merge_cells(start_row=25, start_column=1, end_row=25, end_column=6)
                    
                    # è®¾ç½®å¤‡æ³¨æ–‡å­—çš„æ ·å¼
                    remark_font = Font(size=11)
                    remark_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # æ·»åŠ å¤‡æ³¨å†…å®¹
                    remarks = [
                        "1. å“ç±»æ ¹æ®ä¾›åº”å•†å®é™…é€è´§çš„æƒ…å†µå¡«å†™ï¼Œä¸é€‚ç”¨çš„å¯ç•™ç©º",
                        "2. å‘˜é¤è´§æ¬¾çš„ä¸å«ç¨é‡‘é¢ï¼Œå¦‚é›¶ç¨ç‡ï¼Œé…’åº—éœ€è¦æ ¹æ®å®é™…æ”¶è´§è®°å½•çš„æ€»é‡‘é¢å»æ¢ç®—å«ç¨åŠä¸å«ç¨å¡«å†™",
                        "3. æœ¬å‡½ç”±åŒæ–¹æ ¸å¯¹åŸå§‹æ”¶è´§å•æ®åå¡«å†™ï¼Œä¾›åº”å•†å½“æœˆä¾›è´§æ•°æ®ä¸é…’åº—å½“æœˆåº”ä»˜è´¦æ¬¾é‡‘é¢ä¸€è‡´",
                        "4. ä¾›åº”å•†æ ¹æ®æ ¸å¯¹åç¡®è®¤çš„é‡‘é¢å¼€å…·ç›¸å…³å¢å€¼ç¨å‘ç¥¨ç»™é…’åº—",
                        "5. è¯·ä¾›åº”å•†åœ¨ç¡®è®¤åï¼Œéœ€åŠ ç›–å…¬ç« æˆ–è´¢åŠ¡ä¸“ç”¨ç« ï¼Œæ‰«æåé‚®ä»¶å›ä¼ é…’åº—åšå­˜æ¡£",
                        "6. å»ºè®®éšç¡®è®¤å‡½å‘é€å¢å€¼ç¨å‘ç¥¨å·å’Œå‘ç¥¨é‡‘é¢ä»¥åŠå‘ç¥¨å¤å°ä»¶",
                        "7. ç”µå­é‚®ä»¶å‘é€è‡³ï¼š",
                        "8. æœ¬å‡½è¯·åœ¨æ”¶åˆ°å 2 ä¸ªå·¥ä½œæ—¥å†…è¿”å›",
                        "9. æ‰«æä»¶éœ€æ¸…æ™°æ˜¾ç¤ºï¼šé‡‘é¢ã€ç›–ç« ã€æ—¥æœŸä¸‰è¦ç´ ï¼Œæ¨¡ç³Šæ–‡ä»¶è§†ä¸ºæ— æ•ˆ"
                    ]
                    
                    for i, remark in enumerate(remarks):
                        cell = summary_sheet.cell(row=26+i, column=1, value=remark)
                        cell.font = remark_font
                        cell.alignment = remark_alignment
                        # åˆå¹¶æ¯è¡Œçš„Aè‡³Fåˆ—ï¼Œä½†è·³è¿‡ç¬¬32è¡Œï¼ˆ26+6ï¼‰
                        if 26+i != 32:
                            summary_sheet.merge_cells(start_row=26+i, start_column=1, end_row=26+i, end_column=6)
                    
                    # åœ¨B32å•å…ƒæ ¼ä¸­æ·»åŠ é‚®ç®±åœ°å€
                    email_cell = summary_sheet.cell(row=32, column=2, value=email_address)
                    email_cell.font = remark_font
                    email_cell.alignment = remark_alignment
                    # åˆå¹¶B32åˆ°F32å•å…ƒæ ¼
                    summary_sheet.merge_cells(start_row=32, start_column=2, end_row=32, end_column=6)
                    
                    # åœ¨ç¬¬36è¡ŒAåˆ—æ’å…¥ä¾›åº”å•†ç¡®è®¤æ—¥æœŸæ–‡å­—
                    date_font = Font(size=11)
                    date_alignment = Alignment(horizontal='left', vertical='center')
                    
                    date_cell = summary_sheet.cell(row=36, column=1, value="ä¾›åº”å•†ç¡®è®¤æ—¥æœŸï¼š_______å¹´_______æœˆ_______æ—¥")
                    date_cell.font = date_font
                    date_cell.alignment = date_alignment
                    # åˆå¹¶ä¾›åº”å•†ç¡®è®¤æ—¥æœŸè¡Œçš„Aè‡³Fåˆ—
                    summary_sheet.merge_cells(start_row=36, start_column=1, end_row=36, end_column=6)
                    # åˆå¹¶ç¬¬39è¡Œçš„Aè‡³Fåˆ—
                    summary_sheet.merge_cells(start_row=39, start_column=1, end_row=39, end_column=6)
                    
                    # åœ¨ç¬¬38è¡Œæ’å…¥ä¾›åº”å•†ç›–ç« ç¡®è®¤æ–‡å­—
                    stamp_font = Font(size=13, underline="single")
                    stamp_alignment = Alignment(horizontal='center', vertical='center')
                    
                    stamp_cell = summary_sheet.cell(row=39, column=1, value="ä¾›åº”å•†ç›–ç« ç¡®è®¤")
                    stamp_cell.font = stamp_font
                    stamp_cell.alignment = stamp_alignment
                    # åˆå¹¶ç¬¬39è¡Œçš„Aè‡³Fåˆ—
                    summary_sheet.merge_cells(start_row=39, start_column=1, end_row=39, end_column=6)
                    
                    # è®¾ç½®æ‰€æœ‰æ•°æ®å•å…ƒæ ¼çš„è¾¹æ¡†å’Œå¯¹é½æ–¹å¼
                    from openpyxl.styles import Border, Side
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # è®¾ç½®æ‰€æœ‰å•å…ƒæ ¼çš„è¾¹æ¡†å’Œæ ¼å¼
                    for row in range(14, row_idx + 1):
                        for col in range(1, 5):
                            cell = summary_sheet.cell(row=row, column=col)
                            cell.border = thin_border
                            
                            # ä¸ºæ•°å­—åˆ—è®¾ç½®å¯¹é½æ–¹å¼å’Œæ•°å­—æ ¼å¼
                            if col > 1:  # é‡‘é¢åˆ—
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                                cell.number_format = '#,##0.00'
                            else:  # åˆ†ç±»åˆ—
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # ä¸ºåˆ†ç±»è¡Œæ·»åŠ äº¤æ›¿èƒŒæ™¯è‰²
                    light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    
                    # å‘˜å·¥é¤å…åˆ†ç±»è¡Œ
                    start_row = 5  # å‘˜å·¥é¤å…åˆ†ç±»å¼€å§‹è¡Œ
                    for i, _ in enumerate(ordered_categories):
                        if i % 2 == 1:  # å¶æ•°è¡Œæ·»åŠ æµ…è‰²èƒŒæ™¯
                            for col in range(1, 5):
                                summary_sheet.cell(row=start_row + i, column=col).fill = light_fill
                    
                    # å…¶ä»–é¤å…ï¼ˆè¥ä¸šç‚¹ï¼‰åˆ†ç±»è¡Œ
                    start_row = 5 + len(ordered_categories) + 3  # å…¶ä»–é¤å…ï¼ˆè¥ä¸šç‚¹ï¼‰åˆ†ç±»å¼€å§‹è¡Œ
                    for i, _ in enumerate(ordered_categories):
                        if i % 2 == 1:  # å¶æ•°è¡Œæ·»åŠ æµ…è‰²èƒŒæ™¯
                            for col in range(1, 5):
                                summary_sheet.cell(row=start_row + i, column=col).fill = light_fill
                    
                    # å°†"æ±‡æ€»"sheetæ›´åä¸º"ç¡®è®¤å‡½"
                    summary_sheet.title = "ç¡®è®¤å‡½"
                    
                    # é‡æ–°è®¾ç½®ç¬¬14è¡Œå’Œç¬¬15è¡Œå±…ä¸­å¯¹é½ï¼Œæµ…è“è‰²èƒŒæ™¯è‰²
                    light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    for row in range(14, 16):
                        for col in range(1, 7):
                            cell = summary_sheet.cell(row=row, column=col)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.fill = light_blue_fill
                    
                    # è®¾ç½®ç¬¬2è¡Œåˆ°ç¬¬12è¡Œæ— èƒŒæ™¯è‰²
                    for row in range(2, 13):
                        for col in range(1, 7):
                            cell = summary_sheet.cell(row=row, column=col)
                            cell.fill = PatternFill(fill_type=None)
                    
                    # è®¾ç½®ç¬¬2è¡Œã€ç¬¬5è¡Œã€ç¬¬8è¡Œå’Œç¬¬13è¡Œçš„è¡Œé«˜ä¸º30
                    for row_num in [2, 5, 8, 13]:
                        summary_sheet.row_dimensions[row_num].height = 30
                    
                    # éšè—å“ç±»æ ‡è®°åˆ—ï¼ˆç¬¬14åˆ—ï¼Œå³Nåˆ—ï¼‰
                    ws.column_dimensions['N'].hidden = True
                    
                    # ä¿å­˜æ–‡ä»¶
                    self.log_message(f"  â†’ æ­£åœ¨ä¿å­˜Excelæ–‡ä»¶...")
                    wb.save(output_file)
                    self.log_message(f"  â†’ Excelæ–‡ä»¶ä¿å­˜æˆåŠŸï¼ˆä¿ç•™åŸå§‹æ ¼å¼ï¼‰")
                except Exception as e:
                    self.log_message(f"  â†’ ä¿ç•™æ ¼å¼ä¿å­˜å¤±è´¥ï¼Œå°†ä½¿ç”¨æ ‡å‡†æ–¹å¼ä¿å­˜: {str(e)}")
                    # å¦‚æœä¸Šé¢çš„æ–¹æ³•å¤±è´¥ï¼Œä½¿ç”¨pandasç›´æ¥ä¿å­˜
                    self.log_message(f"  â†’ ä½¿ç”¨æ ‡å‡†æ–¹å¼ä¿å­˜æ–‡ä»¶...")
                    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    if self.edit_in_place_var.get():
                        self.log_message(f"  â†’ å·²ä½¿ç”¨æ ‡å‡†æ–¹å¼ç›´æ¥ä¿®æ”¹åŸæ–‡ä»¶")
                    else:
                        self.log_message(f"  â†’ å·²ä½¿ç”¨æ ‡å‡†æ–¹å¼ä¿å­˜æ–‡ä»¶åˆ°: {output_file}")
            except Exception as e:
                self.log_message(f"  âœ— ä¿å­˜æ–‡ä»¶æ—¶å‡ºé”™: {str(e)}")
                return False
            
            if self.edit_in_place_var.get():
                self.log_message(f"  âœ“ åˆ†ç±»å®Œæˆï¼Œå·²ç›´æ¥ä¿®æ”¹åŸæ–‡ä»¶")
                self.log_message(f"  â†’ æ–‡ä»¶è·¯å¾„: {output_file}")
            else:
                self.log_message(f"  âœ“ åˆ†ç±»å®Œæˆï¼Œæ–‡ä»¶å·²ä¿å­˜")
                self.log_message(f"  â†’ æ–‡ä»¶è·¯å¾„: {output_file}")
            
            # ç»Ÿè®¡å„åˆ†ç±»æ•°é‡å’Œé‡‘é¢
            total_items = len(df)
            
            # æŒ‰è´¢åŠ¡æ ‡è®°åˆ†ç±»ç»Ÿè®¡ï¼ŒæŒ‰æŒ‡å®šé¡ºåºæ˜¾ç¤º
            ordered_categories = ["å¹²è´§", "æµ·é²œ", "é…’ç±»", "é¥®æ–™", "æ°´", "å…¶ä»–"]
            
            # å®šä¹‰å‘˜å·¥é¤å…å’Œå…¶ä»–é¤å…ï¼ˆè¥ä¸šç‚¹ï¼‰
            employee_restaurants = ["å‘˜å·¥é¤å…", "å‘˜å·¥é£Ÿå ‚"]
            
            # åˆå§‹åŒ–æ€»è®¡å˜é‡
            total_employee_untaxed = 0
            total_employee_tax = 0
            total_other_untaxed = 0
            total_other_tax = 0
            
            # å‘˜å·¥é¤å…ç»Ÿè®¡
            employee_items = len(df[df["éƒ¨é—¨"].isin(employee_restaurants)])
            
            for category in ordered_categories:
                # ç­›é€‰è¯¥åˆ†ç±»çš„å‘˜å·¥é¤å…æ•°æ®
                category_df = df[(df[classification_column] == category) & 
                                 (df["éƒ¨é—¨"].isin(employee_restaurants))]
                count = len(category_df)
                
                # è®¡ç®—æœªç¨é‡‘é¢å’Œç¨é¢
                untaxed_amount = category_df["å°è®¡é‡‘é¢(ç»“ç®—)"].sum() if not category_df.empty else 0
                tax_amount = category_df["ç¨é¢(ç»“ç®—)"].sum() if not category_df.empty else 0
                total_amount = untaxed_amount + tax_amount
                
                # æ›´æ–°å‘˜å·¥é¤å…æ€»è®¡
                total_employee_untaxed += untaxed_amount
                total_employee_tax += tax_amount
            
            # å…¶ä»–é¤å…ï¼ˆè¥ä¸šç‚¹ï¼‰ç»Ÿè®¡
            other_items = len(df[~df["éƒ¨é—¨"].isin(employee_restaurants)])
            
            for category in ordered_categories:
                # ç­›é€‰è¯¥åˆ†ç±»çš„å…¶ä»–é¤å…ï¼ˆè¥ä¸šç‚¹ï¼‰æ•°æ®
                category_df = df[(df[classification_column] == category) & 
                                 (~df["éƒ¨é—¨"].isin(employee_restaurants))]
                count = len(category_df)
                
                # è®¡ç®—æœªç¨é‡‘é¢å’Œç¨é¢
                untaxed_amount = category_df["å°è®¡é‡‘é¢(ç»“ç®—)"].sum() if not category_df.empty else 0
                tax_amount = category_df["ç¨é¢(ç»“ç®—)"].sum() if not category_df.empty else 0
                total_amount = untaxed_amount + tax_amount
                
                # æ›´æ–°å…¶ä»–é¤å…ï¼ˆè¥ä¸šç‚¹ï¼‰æ€»è®¡
                total_other_untaxed += untaxed_amount
                total_other_tax += tax_amount
            
            # å¦‚æœæ˜¯æ‰¹å¤„ç†æ¨¡å¼ï¼Œç›´æ¥è¿”å›æˆåŠŸ
            if is_batch:
                return True
            # éæ‰¹å¤„ç†æ¨¡å¼ä¸‹ï¼Œè¯¢é—®ç”¨æˆ·æ˜¯å¦æ‰“å¼€æ–‡ä»¶å¤¹
            message = "æ–‡ä»¶å¤„ç†å®Œæˆï¼Œ" + ("å·²ç›´æ¥ä¿®æ”¹åŸæ–‡ä»¶" if self.edit_in_place_var.get() else f"å·²ä¿å­˜åˆ°:\n{output_file}")
            if messagebox.askyesno("å¤„ç†å®Œæˆ", f"{message}\n\næ˜¯å¦æ‰“å¼€æ–‡ä»¶æ‰€åœ¨æ–‡ä»¶å¤¹ï¼Ÿ"):
                try:
                    # å¦‚æœæ˜¯ç”Ÿæˆç¡®è®¤å‡½ï¼ˆéåŸåœ°ç¼–è¾‘ï¼‰ï¼Œæ‰“å¼€Confirmedæ–‡ä»¶å¤¹
                    if not self.edit_in_place_var.get():
                        output_dir = os.path.join(os.path.dirname(file_path), "Confirmed")
                    else:
                        output_dir = os.path.dirname(output_file)
                    
                    if sys.platform == "win32":
                        os.startfile(output_dir)
                    elif sys.platform == "darwin":  # macOS
                        subprocess.call(["open", output_dir])
                    else:  # Linux
                        subprocess.call(["xdg-open", output_dir])
                except Exception as e:
                    self.log_message(f"æ— æ³•æ‰“å¼€æ–‡ä»¶å¤¹: {str(e)}")
                    messagebox.showerror("é”™è¯¯", f"æ— æ³•æ‰“å¼€æ–‡ä»¶å¤¹:\n{str(e)}")
            
            return True
            
        except Exception as e:
            self.log_message(f"å¤„ç†æ–‡ä»¶æ—¶å‡ºé”™: {str(e)}")
            return False
        finally:
            if not is_batch:
                self.processing = False
                self.process_btn.config(state=NORMAL)
                self.progress['value'] = 100
    
    def bring_to_front(self):
        """å°†çª—å£å¸¦åˆ°å‰å°"""
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after_idle(self.root.attributes, '-topmost', False)
    


if __name__ == "__main__":
    root = Tk()
    app = ProductClassificationApp(root)
    root.mainloop()