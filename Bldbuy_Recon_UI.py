import pandas as pd
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from datetime import datetime
import os
from tkinter import *
from tkinter import filedialog, messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import threading
import shutil
import sys
import subprocess
from Product_Classification_Tool import ProductClassificationApp

# 版本号
VERSION = '2.0.5'

class BldBuyApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"供应商对帐工具集 v{VERSION} - Powered By Cayman Fu @ Sofitel HAIKOU")
        
        # 设置窗口图标
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "favicon.ico")
        if os.path.exists(icon_path):
            self.root.iconbitmap(icon_path)
        
        # 如果不是使用ttk.Window创建的窗口，则设置窗口大小并居中
        if not hasattr(self.root, 'position'):
            self.set_window_geometry(800, 630)
        
        # 创建主题选择下拉框
        self.create_theme_selector()
        
        # 使窗口前台显示
        self.bring_to_front()
        
        # 检查时间验证
        self.check_expiration()
        
        # 检查并确保配置文件存在
        self.ensure_config_file()
            
        # 定义期望的表头字段
        self.expected_headers = [
            "收货日期", "订单号", "商品名称", "实收数量", "基本单位",
            "单价(结算)", "小计金额(结算)", "税额(结算)", "小计价税(结算)", "部门",
            "税率", "供应商/备用金报销账户","商品分类"
        ]
        
        # 创建主框架
        self.main_frame = ttk.Frame(root, padding="10")
        self.main_frame.pack(fill=BOTH, expand=True)
        
        # 创建左右分割的布局
        self.paned_window = ttk.PanedWindow(self.main_frame, orient=HORIZONTAL, bootstyle=PRIMARY)
        self.paned_window.pack(fill=BOTH, expand=True)
        
        # 创建左侧功能按钮区域
        self.left_frame = ttk.Frame(self.paned_window, padding="5")
        self.paned_window.add(self.left_frame, weight=1)
        
        # 创建右侧操作区域
        self.right_frame = ttk.Frame(self.paned_window, padding="5")
        self.paned_window.add(self.right_frame, weight=4)
        
        # 创建左侧功能按钮
        self.create_left_buttons()
        
        # 默认选中"对帐明细表"按钮
        self.handle_button_click(self.show_supplier_panel, ">对帐明细表")
        
        # 初始化状态
        self.processing = False
        
    def set_window_geometry(self, width, height):
        """设置窗口大小并居中"""
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        
    def check_expiration(self):
        """检查时间是否到期"""
        current_date = datetime.now()
        expiration_date = datetime(2025, 12, 31)  # 2025年底到期
        
        if current_date > expiration_date:
            messagebox.showerror("错误", "DLL注册失败，请联系Cayman更新")
            sys.exit(1)
        return True
        
    def get_config_path(self):
        """获取配置文件路径"""
        if getattr(sys, 'frozen', False):
            # 如果是打包后的exe运行
            return os.path.join(os.path.dirname(sys.executable), "config.txt")
        else:
            # 如果是源码运行
            return os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.txt")
    
    def ensure_config_file(self):
        """确保配置文件存在，如果不存在则创建"""
        config_path = self.get_config_path()
        if not os.path.exists(config_path):
            default_config = '''B2:海口索菲特大酒店
D2:海南省海口市龙华区滨海大道105号
E2:符小瑜 0898-31289999
B32:abbyfu@hksft.com
hotelname:海口索菲特大酒店
Sheet_tittle:供货明细表'''
            with open(config_path, 'w', encoding='utf-8') as f:
                f.write(default_config)
            messagebox.showinfo("提示", "已创建默认配置文件：config.txt")
            return True
        return False
        
    def load_theme(self):
        """从配置文件加载主题设置"""
        try:
            config_file = self.get_config_path()
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = dict(line.split(':', 1) for line in f if ':' in line)
                    return config.get('theme', 'sandstone').strip()
        except Exception as e:
            print(f"加载主题设置时出错: {str(e)}")
        return 'sandstone'
    
    def save_theme(self, theme_name):
        """保存主题设置到配置文件"""
        try:
            config_file = self.get_config_path()
            config = {}
            
            # 读取现有配置
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = dict(line.split(':', 1) for line in f if ':' in line)
            
            # 更新主题设置
            config['theme'] = theme_name
            
            # 写入配置文件
            with open(config_file, 'w', encoding='utf-8') as f:
                for key, value in config.items():
                    f.write(f"{key}:{value}\n")
        except Exception as e:
            print(f"保存主题设置时出错: {str(e)}")
    
    def create_theme_selector(self):
        """创建主题选择下拉框"""
        theme_frame = ttk.Frame(self.root)
        theme_frame.pack(side=TOP, fill=X, padx=10, pady=5)
        
        ttk.Label(theme_frame, text="主题：").pack(side=LEFT)
        themes = ttk.Style().theme_names()
        current_theme = self.load_theme()
        ttk.Style().theme_use(current_theme)
        self.theme_var = StringVar(value=current_theme)
        theme_menu = ttk.OptionMenu(
            theme_frame,
            self.theme_var,
            self.theme_var.get(),
            *themes,
            command=self.change_theme,
            bootstyle=SECONDARY
        )
        theme_menu.pack(side=LEFT)
    
    def change_theme(self, theme_name):
        """切换主题"""
        ttk.Style().theme_use(theme_name)
        self.save_theme(theme_name)
    
    def create_left_buttons(self):
        """创建左侧功能按钮"""
        self.left_buttons = []
        buttons = [
            (">对帐明细表", self.show_supplier_panel),
            (">对帐确认函", self.show_confirmation_panel)
        ]
        
        for text, command in buttons:
            btn = ttk.Button(self.left_frame, text=text, command=lambda cmd=command, btn_text=text: self.handle_button_click(cmd, btn_text), width=20, bootstyle=PRIMARY)
            btn.pack(pady=5, padx=5)
            self.left_buttons.append((btn, text))
    
    def handle_button_click(self, command, clicked_text):
        """处理按钮点击事件"""
        # 恢复所有按钮的原始状态
        for btn, text in self.left_buttons:
            btn.configure(text=text, bootstyle=PRIMARY)
        
        # 设置被点击按钮的新状态
        for btn, text in self.left_buttons:
            if text == clicked_text:
                btn.configure(text=text.replace('>', '▶'), bootstyle=SUCCESS)
                break
        
        # 执行原始命令
        command()
    
    def show_supplier_panel(self):
        """显示供应商供货明细面板"""
        self.clear_right_frame()
        self.create_control_panel()
        self.create_log_area()
    
    def show_confirmation_panel(self):
        """显示供应商确认函面板"""
        self.clear_right_frame()
        # 在右侧面板中创建供应商确认函工具实例
        ProductClassificationApp(self.right_frame)
    
    def clear_right_frame(self):
        """清空右侧面板的所有组件"""
        for widget in self.right_frame.winfo_children():
            widget.destroy()
    
    def create_control_panel(self):
        control_frame = ttk.LabelFrame(self.right_frame, text="请选择[收货单商品明细]报表", padding="10", bootstyle=PRIMARY)
        control_frame.pack(fill=X, pady=5)
        
        # 创建文件选择框架
        self.file_frame = ttk.Frame(control_frame)
        self.file_frame.pack(fill=X, pady=5)
        
        # 添加"选择Excel文件"的文本标签
        ttk.Label(self.file_frame, text="选择Excel文件：").pack(side=LEFT, padx=5)
        
        self.input_file_var = StringVar()
        ttk.Entry(self.file_frame, textvariable=self.input_file_var, width=40).pack(side=LEFT, padx=5)
        ttk.Button(self.file_frame, text="浏览...", command=self.select_input_file, bootstyle=SECONDARY).pack(side=LEFT)
        
        # 处理按钮
        self.process_btn = ttk.Button(control_frame, text="开始处理", command=self.start_processing, bootstyle=SUCCESS)
        self.process_btn.pack(pady=10)
        
        # 进度条
        self.progress = ttk.Progressbar(control_frame, orient=HORIZONTAL, mode='determinate', bootstyle=SUCCESS)
        self.progress.pack(fill=X, pady=5)
        
    def create_log_area(self):
        log_frame = ttk.LabelFrame(self.right_frame, text="处理日志", padding="10", bootstyle=PRIMARY)
        log_frame.pack(fill=X, expand=False)
        log_frame.configure(height=200)
        
        self.log_text = Text(log_frame, wrap=WORD, state=DISABLED, height=18)
        self.log_text.pack(fill=X, expand=False)
        
    def select_input_file(self):
        filetypes = [("Excel files", "*.xlsx *.xls")]
        file_paths = filedialog.askopenfilenames(filetypes=filetypes)
        if file_paths:
            self.input_file_var.set("\n".join(file_paths))  # 用换行符分隔多个文件路径
            
    def log_message(self, message):
        """修改后的日志记录方法"""
        # 将消息添加到日志列表
        self.log_messages.append(message)
        
        # 实时显示日志
        self.log_text.config(state=NORMAL)
        if message.startswith("警告："):
            self.log_text.tag_config("warning", foreground="red")
            self.log_text.insert(END, message + "\n", "warning")
        else:
            self.log_text.insert(END, message + "\n")
        self.log_text.see(END)
        self.log_text.config(state=DISABLED)
        
    def start_processing(self):
        if self.processing:
            return
            
        self.processing = True
        self.process_btn.config(state=DISABLED)
        
        # 禁用所有左侧按钮
        for btn, text in self.left_buttons:
            btn.config(state=DISABLED)
        
        self.log_text.delete(1.0, END)
        self.progress['value'] = 0
        
        # 使用线程处理，避免界面卡顿
        threading.Thread(target=self.process_files, daemon=True).start()
        
    def preprocess_excel(self, file_path):
        """预处理Excel文件，自动搜索表头位置"""
        try:
            # 自动搜索表头位置
            header_row = self.find_header_row(file_path)
            
            # 使用找到的表头行读取数据
            df = pd.read_excel(file_path, skiprows=header_row)
            
            # 添加需要保留的退货相关列，排除N-R列数据
            required_columns = self.expected_headers + ['退货', '合计退货数量', '退货合计金额(结算)', '退货合计税额(结算)', '退货合计价税(结算)']
            
            # 过滤并重新排列列，排除N-R列的数据
            exclude_columns = df.iloc[:, 13:17].columns.tolist()  # N-R列的索引是13-17
            df = df.drop(columns=exclude_columns, errors='ignore')
            
            # 过滤并保留所需列
            df_filtered = df.reindex(columns=[col if col != '单位' else '基本单位' for col in required_columns if col in df.columns or col == '基本单位'])
            
            # 检查是否找到了必要的列
            missing_columns = [col for col in self.expected_headers if col not in df_filtered.columns]
            if missing_columns:
                self.log_message(f"警告：文件中缺少必要的列：{', '.join(missing_columns)}，请检查是否选择了正确的文件")
                return None
            
            # 处理收货日期，去掉时间部分
            if '收货日期' in df_filtered.columns:
                df_filtered['收货日期'] = pd.to_datetime(df_filtered['收货日期'], errors='coerce').dt.strftime('%Y-%m-%d')
            
            return df_filtered.dropna(how='all')
        except Exception as e:
            self.log_message(f"警告：处理文件时出错，请检查是否选择了正确的文件。错误信息：{str(e)}")
            return None
        
    def find_header_row(self, file_path):
        """自动搜索Excel文件中的表头行"""
        # 最大搜索行数
        max_rows = 50
        
        # 读取前max_rows行来查找表头
        sample_df = pd.read_excel(file_path, nrows=max_rows, header=None)
        
        # 定义匹配度阈值（至少需要匹配的表头数量）
        min_match_threshold = 3
        
        # 遍历每一行，检查是否包含足够多的预期表头
        for i in range(max_rows):
            row = sample_df.iloc[i].astype(str)
            # 计算当前行与预期表头的匹配数量
            matches = sum(1 for header in self.expected_headers if any(header in str(cell) for cell in row))
            
            # 如果匹配数量超过阈值，认为找到了表头行
            if matches >= min_match_threshold:
                self.log_message(f"找到表头行: 第{i+1}行，匹配度: {matches}/{len(self.expected_headers)}")
                return i
        
        # 如果没有找到，使用默认值
        self.log_message("未找到表头行，使用默认值(35)")
        return 35
        
    def process_files(self):
        try:
            # 初始化日志列表和文件夹
            self.log_messages = []
            folders = {
                'output': "export",
                'archive': "archive"
            }
            
            # 获取输入文件列表
            input_files = [f for f in self.input_file_var.get().split("\n") if f]
            if not input_files:
                self.log_message("请先选择要处理的Excel文件")
                return
            
            # 创建必要的文件夹
            for folder in folders.values():
                os.makedirs(folder, exist_ok=True)
            
            # 获取配置信息
            header_rows = self.get_config_header_rows()
            
            # 批量处理文件
            total_files = len(input_files)
            for index, input_file in enumerate(input_files, 1):
                try:
                    self.log_message(f"\n正在处理文件: {os.path.basename(input_file)}")
                    
                    # 预处理数据
                    df_filtered = self.preprocess_excel(input_file)
                    if df_filtered is None:  # 预处理失败
                        continue
                    
                    # 获取年月信息
                    year_month = self.get_year_month(df_filtered)
                    if not year_month:
                        self.log_message("错误：无法从文件中获取有效的收货日期，请检查是否选择了正确的文件。")
                        continue
                    
                    # 创建年月文件夹
                    year_month_folder = os.path.join(folders['output'], year_month)
                    os.makedirs(year_month_folder, exist_ok=True)
                    
                    # 分组并处理数据
                    self.process_grouped_data(df_filtered, year_month, year_month_folder, header_rows)
                    
                    # 归档文件
                    self.archive_file(input_file, folders['archive'])
                    
                    # 更新进度
                    self.update_progress(index, total_files)
                    
                except Exception as e:
                    self.log_message(f"处理文件 {os.path.basename(input_file)} 时出错: {str(e)}")
            
            # 显示处理结果
            self.show_processing_results(folders['output'])
            
        except Exception as e:
            self.log_message(f"处理过程中发生错误: {str(e)}")
        finally:
            self.processing = False
            self.process_btn.config(state=NORMAL)
            
            # 重新启用所有左侧按钮
            for btn, text in self.left_buttons:
                btn.config(state=NORMAL)
            
    def get_config_header_rows(self):
        """获取配置文件中的标题信息"""
        try:
            application_path = os.path.dirname(sys.executable if getattr(sys, 'frozen', False) 
                                              else os.path.abspath(__file__))
            config_file = os.path.join(application_path, 'config.txt')
            
            if not os.path.exists(config_file):
                self.log_message("警告：未找到config.txt文件,将会导致对帐单标题错误")
                return []
            
            with open(config_file, 'r', encoding='utf-8') as f:
                config = dict(line.split(':', 1) for line in f if ':' in line)
            
            hotelname = config.get('hotelname', '').strip()
            sheet_title = config.get('Sheet_tittle', '').strip()
            
            return [
                [''] * 5 + [hotelname] + [''] * 7,
                [''] * 5 + [sheet_title] + [''] * 7,
                [''] * 13,
                [''] * 13,
                [''] * 13
            ]
        except Exception as e:
            self.log_message(f"读取配置文件时出错: {str(e)}")
            return []
            
    def get_year_month(self, df):
        """从数据中获取年月信息"""
        if '收货日期' not in df.columns:
            self.log_message("错误：文件中缺少收货日期列，请检查是否选择了正确的文件。")
            return None
            
        try:
            # 检查收货日期列是否有数据
            if df['收货日期'].empty or df['收货日期'].isna().all():
                self.log_message("错误：文件中的收货日期列没有任何数据，请检查是否选择了正确的文件。")
                return None
                
            earliest_date = df['收货日期'].min()
            if not earliest_date or pd.isna(earliest_date):
                self.log_message("错误：文件中没有有效的收货日期，请检查是否选择了正确的文件。")
                return None
                
            try:
                return datetime.strptime(earliest_date, '%Y-%m-%d').strftime('%Y-%m')
            except ValueError:
                self.log_message("错误：收货日期格式不正确，请检查是否选择了正确的文件。")
                return None
        except Exception as e:
            self.log_message(f"错误：处理收货日期时出错，请检查是否选择了正确的文件。错误信息：{str(e)}")
            return None
            
    def process_grouped_data(self, df, year_month, year_month_folder, header_rows):
        """处理分组数据"""
        # 排序和分组
        sort_columns = ['收货日期', '税率']
        if all(col in df.columns for col in sort_columns):
            df = df.sort_values(by=sort_columns)
        else:
            self.log_message("警告：文件中缺少排序所需的列，将不按顺序处理数据。")
        
        # 使用向量化操作处理分组
        for group_name, group_data in df.groupby(['供应商/备用金报销账户'], as_index=False):
            self.process_group_data(group_name, group_data, year_month, year_month_folder, header_rows)
            
    def archive_file(self, input_file, archive_folder):
        """归档处理完的文件"""
        try:
            base_name = os.path.basename(input_file)
            archive_path = os.path.join(archive_folder, base_name)
            
            # 如果文件已存在，添加时间戳
            if os.path.exists(archive_path):
                base, ext = os.path.splitext(base_name)
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                archive_path = os.path.join(archive_folder, f"{base}_{timestamp}{ext}")
            
            shutil.move(input_file, archive_path)
            self.log_message(f"已成功归档文件 {base_name}")
        except Exception as e:
            self.log_message(f"归档文件时出错: {str(e)}")
            
    def update_progress(self, current, total):
        """更新进度条"""
        progress_value = int((current / total) * 100)
        self.progress['value'] = progress_value
        self.root.update_idletasks()
        
    def show_processing_results(self, output_folder):
        """显示处理结果和警告信息"""
        # 显示警告和错误信息
        error_messages = [msg for msg in self.log_messages if msg.startswith("错误：")]
        warning_messages = [msg for msg in self.log_messages if msg.startswith("警告：")]
        
        if error_messages:
            self.log_message("\n处理过程中出现以下错误：")
            self.log_text.config(state=NORMAL)
            for msg in error_messages:
                self.log_text.insert(END, msg + "\n", "warning")
            self.log_text.see(END)
            self.log_text.config(state=DISABLED)
            
            # 如果有错误，不显示成功完成的消息
            self.progress['value'] = 100
            return
            
        if warning_messages:
            self.log_message("\n所有文件处理完成。以下是处理过程中的警告信息：")
            self.log_text.config(state=NORMAL)
            for msg in warning_messages:
                self.log_text.insert(END, msg + "\n", "warning")
            self.log_text.see(END)
            self.log_text.config(state=DISABLED)
        else:
            self.log_message("\n所有文件处理完成，没有发现警告信息。")
        
        self.progress['value'] = 100
        
        # 只有在没有错误时才询问是否打开输出目录
        if messagebox.askyesno("处理完成", "所有文件处理已完成，是否打开输出文件夹？"):
            try:
                if sys.platform == "darwin":  # macOS
                    subprocess.call(["open", output_folder])
                elif sys.platform == "win32":  # Windows
                    os.startfile(output_folder)
                else:  # Linux
                    subprocess.call(["xdg-open", output_folder])
            except:
                self.log_message("无法打开文件夹，请手动访问：")
                self.log_message(output_folder)
            
    def process_group_data(self, group_name, group_data, year_month, year_month_folder, header_rows):
        """处理每个分组的数据"""
        supplier_account = group_name
        
        try:
            # 预处理数据
            df_processed, output_filepath = self.prepare_group_data(group_name, group_data, year_month, year_month_folder)
            
            # 创建工作簿和工作表
            wb = Workbook()
            ws = wb.active
            ws.title = "Statement"
            
            # 写入数据
            self.write_excel_content(ws, df_processed, group_data, header_rows)
            
            # 设置样式并保存
            self.apply_styles(ws)
            wb.save(output_filepath)
            
            self.log_message(f"已成功创建 {os.path.basename(output_filepath)}")
            
        except Exception as e:
            self.log_message(f"处理供应商 {supplier_account} 的数据时出错: {str(e)}")
    
    def prepare_group_data(self, group_name, group_data, year_month, year_month_folder):
        """准备分组数据"""
        # 检查跨月
        group_data['收货日期'] = pd.to_datetime(group_data['收货日期'], errors='coerce').dt.strftime('%Y-%m-%d')
        unique_months = pd.to_datetime(group_data['收货日期']).dt.strftime('%Y-%m').unique()
        if len(unique_months) > 1:
            self.log_message(f"警告：供应商 {group_name} 的收货日期包含跨月数据，请核查。包含的月份有：{', '.join(unique_months)}")
        
        # 预处理数据
        df_processed = group_data.reindex(columns=self.expected_headers).fillna('')
        df_processed['税率'] = df_processed['税率'].apply(lambda x: f"{int(float(x) * 100)}%" if pd.notna(x) else '0%')
        
        # 处理退货数据：将退货金额转换为负数
        if '退货' in group_data.columns:
            return_mask = (group_data['退货'] == '是')
            amount_columns = ['单价(结算)', '小计金额(结算)', '税额(结算)', '小计价税(结算)']
            
            for col in amount_columns:
                if col in df_processed.columns:
                    # 将退货行的金额列转换为负数
                    df_processed.loc[return_mask, col] = df_processed.loc[return_mask, col].apply(
                        lambda x: -abs(float(x)) if pd.notna(x) and str(x).replace('.', '').replace('-', '').isdigit() else x
                    )
        
        # 构建文件路径
        sanitized_name = ''.join([c if c.isalnum() or c in (' ', '.') else '_' for c in str(group_name)]).strip('_')
        output_filepath = os.path.join(year_month_folder, f"{year_month}_{sanitized_name}.xlsx")
        
        return df_processed, output_filepath
    
    def write_excel_content(self, ws, df_processed, group_data, header_rows):
        """写入Excel内容"""
        # 写入表头
        for row in header_rows + [self.expected_headers]:
            ws.append(row)
        
        # 写入数据行（包括已处理的退货负数数据）
        for i, row in enumerate(df_processed.values.tolist()):
            ws.append(row)
        
        # 添加合计行
        self.add_total_row(ws, df_processed)
        
        # 填入固定文字
        ws['A3'] = '供应商名称：'
        ws['A4'] = '对帐周期：'
        ws['C3'] = '小计金额(结算)：'
        ws['C4'] = '税额(结算)：'
        ws['C5'] = '小计价税(结算)：'
        
        # 填入数据
        supplier_name = df_processed['供应商/备用金报销账户'].iloc[0] if not df_processed.empty else ''
        ws['B3'] = supplier_name
        
        # 提取年月
        dates = pd.to_datetime(df_processed['收货日期'])
        if not dates.empty:
            year_month = dates.min().strftime('%Y年%m月')
            ws['B4'] = year_month
        
        # 计算包含退货负数的合计数据
        total_subtotal = df_processed['小计金额(结算)'].astype(float).sum()
        total_tax = df_processed['税额(结算)'].astype(float).sum()
        total_amount = df_processed['小计价税(结算)'].astype(float).sum()
        
        # 如果有退货数据，需要额外计算退货部分的负数
        if '退货' in group_data.columns:
            return_mask = (group_data['退货'] == '是')
            if return_mask.any():
                return_data = group_data[return_mask]
                # 退货金额已经在 prepare_group_data 中转换为负数，这里直接使用
                pass
        
        # 填入合计数据
        ws['D3'] = '{:.2f}'.format(total_subtotal)
        ws['D4'] = '{:.2f}'.format(total_tax)
        ws['D5'] = '{:.2f}'.format(total_amount)
        
        # 设置单元格对齐方式
        # A3A4右对齐
        ws['A3'].alignment = Alignment(horizontal="right", vertical="center")
        ws['A4'].alignment = Alignment(horizontal="right", vertical="center")
        
        # B3B4左对齐
        ws['B3'].alignment = Alignment(horizontal="left", vertical="center")
        ws['B4'].alignment = Alignment(horizontal="left", vertical="center")
        
        # C3C4C5右对齐
        ws['C3'].alignment = Alignment(horizontal="right", vertical="center")
        ws['C4'].alignment = Alignment(horizontal="right", vertical="center")
        ws['C5'].alignment = Alignment(horizontal="right", vertical="center")
        
        # D3D4D5左对齐
        ws['D3'].alignment = Alignment(horizontal="left", vertical="center")
        ws['D4'].alignment = Alignment(horizontal="left", vertical="center")
        ws['D5'].alignment = Alignment(horizontal="left", vertical="center")
        
        # 隐藏L列和M列
        ws.column_dimensions['L'].hidden = True
        ws.column_dimensions['M'].hidden = True
    

    
    def add_total_row(self, ws, df_processed):
        """添加合计行"""
        subtotal_amount = df_processed['小计金额(结算)'].astype(float).sum()
        tax_amount = df_processed['税额(结算)'].astype(float).sum()
        total_amount = df_processed['小计价税(结算)'].astype(float).sum()
        
        last_row = ws.max_row + 1
        totals = {
            "单价(结算)": "合计",
            "小计金额(结算)": "{:.2f}".format(subtotal_amount),
            "税额(结算)": "{:.2f}".format(tax_amount),
            "小计价税(结算)": "{:.2f}".format(total_amount)
        }
        
        # 会计专用格式
        accounting_format = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'
        
        for col, value in totals.items():
            cell = ws.cell(row=last_row, column=self.expected_headers.index(col) + 1, value=value)
            # 为数字列应用会计格式
            if col in ["小计金额(结算)", "税额(结算)", "小计价税(结算)"]:
                cell.number_format = accounting_format
                # 设置右对齐
                cell.alignment = Alignment(horizontal='right', vertical='center')
        
    def apply_styles(self, ws):
        """应用样式到工作表"""
        from openpyxl.styles import Border, Side
        import gc
        
        # 使用缓存和批量处理优化性能
        styles_cache = self._create_styles_cache()
        
        # 1. 批量设置页面属性
        self._apply_page_settings(ws)
        
        # 2. 批量设置列宽（使用缓存）
        self._apply_column_widths(ws)
        
        # 3. 批量应用单元格样式（使用缓存和生成器）
        self._apply_cell_styles_optimized(ws, styles_cache)
        
        # 4. 清理缓存和内存
        styles_cache.clear()
        gc.collect()
        
    def _create_styles_cache(self):
        """创建并缓存常用样式"""
        return {
            'header': {
                'fill': None,  # 移除背景色
                'font': Font(color='000000', size=13, name='微软雅黑', bold=False),  # 取消加粗
                'border': None,  # 移除所有边框
                'alignment': Alignment(horizontal="center", vertical="center")
            },
            'data': {
                'font': Font(size=13, name='微软雅黑'),
                'border': Border(left=Side(style='hair', color='D3D3D3'), right=Side(style='hair', color='D3D3D3'),
                                top=Side(style='hair', color='D3D3D3'), bottom=Side(style='hair', color='D3D3D3')),
                'alignment': Alignment(horizontal="center", vertical="center")
            },
            'even_row': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        }
        
    def _apply_page_settings(self, ws):
        """批量应用页面设置"""
        # 页面属性批量设置
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.horizontalCentered = True
        ws.page_setup.verticalCentered = False
        ws.sheet_view.zoomScale = 80  # 设置页面缩放比例为80%
            
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = '1:6'
        ws.freeze_panes = 'A7'
        
        # 页边距批量设置
        margins = {'left': 0.31, 'right': 0.31, 'top': 0.31, 'bottom': 0.79, 'header': 0.31, 'footer': 0.31}
        ws.page_margins = PageMargins(**{k: v * 0.3937 for k, v in margins.items()})
        
        # 页脚设置
        ws.oddFooter.center.font = '微软雅黑'
        ws.oddFooter.center.size = 11
        ws.oddFooter.center.text = "Page &[Page] of &[Pages]"
        
    def _apply_column_widths(self, ws):
        """批量应用列宽设置（使用缓存）"""
        widths = {
            '订单号': 35, '收货日期': 18, '商品名称': 22, '实收数量': 12, '基本单位': 12,
            '单价(结算)': 24, '小计金额(结算)': 24, '税额(结算)': 20, '小计价税(结算)': 20,
            '部门': 20, '供应商/备用金报销账户': 36, '商品分类': 24
        }
        
        # 使用生成器表达式优化内存使用
        col_map = ((header, ws.cell(row=1, column=idx).column_letter)
                   for idx, header in enumerate(self.expected_headers, 1))
        
        # 批量设置列宽
        for header, col_letter in col_map:
            if header in widths:
                ws.column_dimensions[col_letter].width = widths[header]
                
    def _apply_cell_styles_optimized(self, ws, styles_cache):
        """优化的单元格样式应用（使用缓存和生成器）"""
        max_col = len(self.expected_headers)
        
        # 使用生成器优化内存使用
        def cell_generator():
            for row in ws.iter_rows(min_row=1, max_col=max_col, max_row=ws.max_row):
                yield row
        
        # 批量应用样式
        for row in cell_generator():
            row_num = row[0].row
            
            # 检查行中是否有负数（用于设置整行黄色背景）
            has_negative = False
            if row_num > 6 and row_num < ws.max_row:  # 只检查数据行，不包括表头和合计行
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        has_negative = True
                        break
            
            # 确定行样式
            if row_num <= 6:  # 前6行（包括第6行表头）
                style = styles_cache['header']
            elif row_num == 7:  # 第7行（数据开始行）
                style = {
                    'fill': None,
                    'font': Font(color='000000', size=13, name='微软雅黑', bold=False),
                    'border': None,  # 取消第七行下框线
                    'alignment': Alignment(horizontal="center", vertical="center")
                }
            elif row_num == ws.max_row:  # 合计行
                style = {
                    'fill': None,
                    'font': Font(color='000000', size=13, name='微软雅黑', bold=True),
                    'border': Border(top=Side(style='thin', color='000000')),
                    'alignment': Alignment(horizontal="center", vertical="center"),
                    'preserve_format': True  # 标记保留数字格式
                }
            else:
                style = styles_cache['data']
                
            # 批量应用样式
            for cell in row:
                if style == styles_cache['header']:
                    # 对于表头行（1-6行），不应用背景色
                    if style['fill'] is not None:
                        cell.fill = style['fill']
                    cell.font = style['font']
                    
                    # 为第二行设置下边框
                    if row_num == 2:
                        cell.border = Border(bottom=Side(style='thin', color='000000'))
                    # 为第六行（表头行）设置上边框和下边框
                    elif row_num == 6:
                        cell.border = Border(
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thin', color='000000')
                        )
                    else:
                        cell.border = style['border']
                    
                    # 为第3-5行设置右对齐
                    if row_num in [3, 4, 5]:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    # 为表头行的F列到I列设置右对齐
                    elif row_num == 6 and cell.column in [6, 7, 8, 9]:  # F列到I列（单价、小计金额、税额、小计价税）
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = style['alignment']
                    
                    # 设置第一行和第二行的行高
                    if row_num <= 2:
                        ws.row_dimensions[row_num].height = 26
                    # 设置第七行的行高
                    elif row_num == 6:
                        ws.row_dimensions[row_num].height = 30
                else:
                    cell.font = style['font']
                    cell.border = style['border']
                    
                    # 对于合计行的数字列，保持会计格式和右对齐
                    if row_num == ws.max_row and cell.column in [7, 8, 9]:  # 合计行的小计金额、税额、小计价税列
                        # 保持add_total_row中设置的会计格式和右对齐，不覆盖
                        pass
                    else:
                        cell.alignment = style['alignment']
                    
                    # 设置背景色：如果行中有负数则设置黄色背景，否则按奇偶行设置
                    if has_negative:
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    elif row_num % 2 == 0:
                        cell.fill = styles_cache['even_row']
                        
                    # 负数字体设置为红色
                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.font = Font(size=13, name='微软雅黑', color='FF0000')
                    
    def bring_to_front(self):
        """将窗口带到前台"""
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after_idle(self.root.attributes, '-topmost', False)
        
if __name__ == "__main__":
    root = ttk.Window(
        title=f"供应商对帐工具 v{VERSION} - Powered By Cayman Fu @ Sofitel HAIKOU",
        themename="cosmo",
        size=(1024, 768),
        position=None,  # 居中显示
        minsize=(1024, 768),
        resizable=(True, True),
    )
    app = BldBuyApp(root)
    root.mainloop()
